import os
import sys
import ctypes

# Set Windows taskbar icon (must be done before creating Tk window)
if sys.platform == 'win32':
    try:
        # Tell Windows this is a unique app (not grouped with Python apps)
        myappid = 'WikimediaBangladesh.PyPan.BatchUploader.0.2.1'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except Exception:
        pass

# Detect if running as compiled executable (Nuitka) or as script
if getattr(sys.modules[__name__], '__compiled__', False):
    # Running as Nuitka compiled executable
    CONFIG_DIR = os.path.join(os.path.expanduser('~'), '.pypan')
    os.makedirs(CONFIG_DIR, exist_ok=True)
    # For Nuitka, pywikibot files are in the same directory as the executable
    PYWIKIBOT_DATA_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    # Running as script
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    CONFIG_DIR = SCRIPT_DIR
    PYWIKIBOT_DATA_DIR = CONFIG_DIR

USER_CONFIG_PATH = os.path.join(CONFIG_DIR, 'user-config.py')
PASSWORD_FILE_PATH = os.path.join(CONFIG_DIR, 'user-password.py')

import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import pandas as pd
import json
try:
    import openpyxl
except ImportError:
    openpyxl = None
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
import logging
from datetime import datetime
import shutil
from PIL import Image
from urllib.parse import urlparse
import tempfile

try:
    from moviepy import VideoFileClip
    MOVIEPY_AVAILABLE = True
except ImportError:
    MOVIEPY_AVAILABLE = False

try:
    import yt_dlp
    YT_DLP_AVAILABLE = True
except ImportError:
    YT_DLP_AVAILABLE = False
    
# Fix for bundled resources in compiled executable
def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller/Nuitka"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        if getattr(sys.modules[__name__], '__compiled__', False):
            # Nuitka compiled
            base_path = os.path.dirname(os.path.abspath(sys.executable))
        else:
            # Running as script
            base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

ALLOWED_EXTENSIONS = {
    '.svg', '.png', '.jpg', '.jpeg', '.gif', '.tiff', '.tif', '.webp', '.xcf',
    '.oga', '.ogg', '.mid', '.midi', '.wav', '.flac', '.mp3', '.opus',
    '.webm', '.ogv', '.ogx',
    '.apng',
    '.pdf', '.djvu',
    '.stl'
}

# Popular video formats that need conversion to WebM
VIDEO_FORMATS_TO_CONVERT = {
    '.mp4', '.avi', '.mov', '.mkv', '.flv', '.wmv', '.m4v', '.mpeg', '.mpg', '.3gp', '.m2v'
}

def safe_chmod(path, mode):
    try:
        os.chmod(path, mode)
    except PermissionError:
        pass
    except Exception as e:
        print(f"Warning: could not chmod {path}: {e}")

def sanitize_filename(filename):
    import re
    illegal_chars = r'[:#<>\[\]|{}/\\]'
    sanitized = re.sub(illegal_chars, '-', filename)
    sanitized = re.sub(r'~{3,}', '-', sanitized)
    sanitized = re.sub(r'[\x00-\x1F\x7F]', '-', sanitized)
    # Replace multiple consecutive hyphens with a single hyphen
    sanitized = re.sub(r'-+', '-', sanitized)
    # Remove leading/trailing hyphens and spaces
    sanitized = sanitized.strip('- ')
    
    return sanitized

class PyPan:
    def __init__(self, root):
        self.root = root
        self.root.title("PyPan")
        self.root.geometry("800x600")
        
        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar(value="upload_results.xlsx")
        self.num_workers_var = tk.IntVar(value=1)
        self.username = None
        self.password = None
        self.is_logged_in = False
        self.is_running = False
        self.is_paused = False
        self.total_files = 0
        self.processed_files = 0
        self.successful_uploads = 0
        self.failed_uploads = 0
        self.start_time = None
        self.executor = None
        self.site = None
        self.internet_status = tk.StringVar(value="Unknown")
        
        self.results = []
        self.results_lock = threading.Lock()
        self.stop_event = threading.Event()
        
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[logging.StreamHandler()]
        )
        self.logger = logging.getLogger(__name__)
               
        self.setup_ui()
        self.check_external_dependencies()
        # Check internet connection on startup
        threading.Thread(target=self.test_internet_connection, daemon=True).start()
        
    def check_external_dependencies(self):
        """Check for external dependencies and warn user"""
        missing = []
        
        # Check for yt-dlp 
        if not YT_DLP_AVAILABLE:
            missing.append("yt-dlp (for YouTube downloads)")
        
        # Check for moviepy 
        if not MOVIEPY_AVAILABLE:
            missing.append("moviepy (for video conversion)")
        
        if missing:
            msg = "Optional dependencies not found:\n\n" + "\n".join(f"• {m}" for m in missing)
            msg += "\n\nSome features may not work without these libraries."
            msg += "\nInstall via: pip install yt-dlp moviepy"
            self.log_message(msg, "WARNING")    
    
    def cleanup_config_files(self):
        """Delete config files and extra pywikibot artifacts in CONFIG_DIR"""
        try:
            # removing user-config and password files
            for p in (USER_CONFIG_PATH, PASSWORD_FILE_PATH):
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except Exception as e:
                    print(f"Warning: could not remove {p}: {e}")

            # removing apicache directory
            apicache_dir = os.path.join(CONFIG_DIR, 'apicache')
            try:
                if os.path.isdir(apicache_dir):
                    shutil.rmtree(apicache_dir, ignore_errors=True)
            except Exception as e:
                print(f"Warning: could not remove apicache dir {apicache_dir}: {e}")

            # removing throttle control file
            throttle_path = os.path.join(CONFIG_DIR, 'throttle.ctrl')
            try:
                if os.path.exists(throttle_path):
                    os.remove(throttle_path)
            except Exception as e:
                print(f"Warning: could not remove {throttle_path}: {e}")

            # removing lwp file
            try:
                uname = getattr(self, 'username', '') or ''
                lwp_name = f"pywikibot-{uname.replace(' ', '_')}.lwp"
                lwp_path = os.path.join(CONFIG_DIR, lwp_name)
                if os.path.exists(lwp_path):
                    os.remove(lwp_path)
            except Exception as e:
                print(f"Warning: could not remove lwp file: {e}")

            # removing upload_log.txt
            try:
                log_path = os.path.join(CONFIG_DIR, 'upload_log.txt')
                if os.path.exists(log_path):
                    os.remove(log_path)
            except Exception as e:
                print(f"Warning: could not remove upload_log.txt: {e}")

        except Exception as e:
            print(f"Warning: cleanup_config_files failed: {e}")
        
    def setup_ui(self):
        # Set window icon
        try:
            # When compiled, use the exe's embedded icon
            if getattr(sys.modules[__name__], '__compiled__', False):
                self.root.iconbitmap(default=sys.executable)
            else:
                # When running as script, use icon file
                icon_path = get_resource_path('icon.ico')
                if os.path.exists(icon_path):
                    self.root.iconbitmap(icon_path)
        except Exception as e:
            # Icon is optional, continue without it
            pass
        
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        config_frame = ttk.LabelFrame(main_frame, text="Configuration", padding="10")
        config_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        ttk.Label(config_frame, text="Login Status:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.login_status_var = tk.StringVar(value="Not logged in")
        self.login_status_label = ttk.Label(config_frame, textvariable=self.login_status_var, font=('Arial', 10, 'bold'), foreground='red')
        self.login_status_label.grid(row=0, column=1, sticky=tk.W, padx=(0, 5))
        
        self.login_btn = ttk.Button(config_frame, text="Login", command=self.show_login_window)
        self.login_btn.grid(row=0, column=2, sticky=tk.W)
        
        # Family and Language settings 
        ttk.Label(config_frame, text="Family:").grid(row=0, column=3, sticky=tk.W, padx=(10, 5))
        self.family_var = tk.StringVar(value="commons")
        ttk.Entry(config_frame, textvariable=self.family_var, width=12).grid(row=0, column=4, sticky=tk.W, padx=(0, 5))
        
        ttk.Label(config_frame, text="Lang:").grid(row=0, column=5, sticky=tk.W, padx=(10, 5))
        self.mylang_var = tk.StringVar(value="commons")
        ttk.Entry(config_frame, textvariable=self.mylang_var, width=12).grid(row=0, column=6, sticky=tk.W, padx=(0, 10))
        
        ttk.Label(config_frame, text="Input File:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5))
        ttk.Entry(config_frame, textvariable=self.input_file, width=50).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 5))
        ttk.Button(config_frame, text="Browse", command=self.browse_input_file).grid(row=1, column=2, sticky=tk.W)
        
        ttk.Label(config_frame, text="Output File:").grid(row=2, column=0, sticky=tk.W, padx=(0, 5))
        ttk.Entry(config_frame, textvariable=self.output_file, width=30).grid(row=2, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Label(config_frame, text="Parallelization:").grid(row=4, column=0, sticky=tk.W, padx=(0,5))
        self.num_workers_var = tk.IntVar(value=1)
        ttk.Entry(config_frame, textvariable=self.num_workers_var, width=5).grid(row=4, column=1, sticky=tk.W, padx=(0,10))
        
        ttk.Label(config_frame, text="Pause Between Retry(s):").grid(row=4, column=2, sticky=tk.W, padx=(10,5))
        self.pause_seconds_var = tk.IntVar(value=10)
        ttk.Entry(config_frame, textvariable=self.pause_seconds_var, width=5).grid(row=4, column=3, sticky=tk.W, padx=(0,10))
        
        ttk.Label(config_frame, text="Max Retry Attempts:").grid(row=4, column=4, sticky=tk.W, padx=(10,5))
        self.max_attempts_var = tk.IntVar(value=10)
        ttk.Entry(config_frame, textvariable=self.max_attempts_var, width=5).grid(row=4, column=5, sticky=tk.W, padx=(0,10))
        
        ttk.Label(config_frame, text="Pause After Upload(s):").grid(row=5, column=2, sticky=tk.W, padx=(10,5))
        self.pause_after_upload_var = tk.DoubleVar(value=0.2)
        ttk.Entry(config_frame, textvariable=self.pause_after_upload_var, width=5).grid(row=5, column=3, sticky=tk.W, padx=(0,10))
        
        ttk.Label(config_frame, text="Ignore Warnings:").grid(row=5, column=0, sticky=tk.W, padx=(0,5))
        self.ignore_warnings_var = tk.StringVar(value="True")
        ignore_dropdown = ttk.Combobox(config_frame, textvariable=self.ignore_warnings_var, values=["True", "False"], width=5, state="readonly")
        ignore_dropdown.grid(row=5, column=1, sticky=tk.W, padx=(0,10))

        ttk.Label(config_frame, text="Internet Status:").grid(row=3, column=0, sticky=tk.W, padx=(0, 5))
        self.internet_status_label = ttk.Label(config_frame, textvariable=self.internet_status, foreground="gray")
        self.internet_status_label.grid(row=3, column=1, sticky=tk.W, padx=(0, 10))
        
        config_frame.columnconfigure(1, weight=1)
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.start_btn = ttk.Button(button_frame, text="Start Upload", command=self.start_upload)
        self.start_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.pause_btn = ttk.Button(button_frame, text="Pause", command=self.pause_upload, state=tk.DISABLED)
        self.pause_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.stop_btn = ttk.Button(button_frame, text="Stop", command=self.stop_upload, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.test_connection_btn = ttk.Button(button_frame, text="Test Connection", command=lambda: self.test_internet_connection(show_success=True))
        self.test_connection_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.clear_btn = ttk.Button(button_frame, text="Reset", command=self.clear_reset)
        self.clear_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        self.status_label = ttk.Label(progress_frame, text="Ready")
        self.status_label.grid(row=1, column=0, sticky=tk.W)
        
        self.stats_label = ttk.Label(progress_frame, text="Files: 0/0 | Success: 0 | Failed: 0")
        self.stats_label.grid(row=2, column=0, sticky=tk.W)
        
        self.time_label = ttk.Label(progress_frame, text="Time: 00:00:00 | ETA: --:--:--")
        self.time_label.grid(row=3, column=0, sticky=tk.W)
        
        progress_frame.columnconfigure(0, weight=1)
        
        log_frame = ttk.LabelFrame(main_frame, text="Log", padding="10")
        log_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        self.log_text = tk.Text(log_frame, height=15, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
    def show_login_window(self):
        """Show login dialog"""
        login_window = tk.Toplevel(self.root)
        login_window.title("Wikimedia - Login")
        login_window.geometry("400x275")
        login_window.resizable(False, False)
        login_window.transient(self.root)
        login_window.grab_set()
        
        # Center window
        login_window.update_idletasks()
        width = login_window.winfo_width()
        height = login_window.winfo_height()
        x = (login_window.winfo_screenwidth() // 2) - (width // 2)
        y = (login_window.winfo_screenheight() // 2) - (height // 2)
        login_window.geometry(f'{width}x{height}+{x}+{y}')
        
        main_frame = ttk.Frame(login_window, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        title_label = ttk.Label(main_frame, text="Wikimedia Login", 
                               font=('Arial', 14, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        ttk.Label(main_frame, text="Username:").grid(row=1, column=0, sticky=tk.W, pady=5)
        username_var = tk.StringVar()
        username_entry = ttk.Entry(main_frame, textvariable=username_var, width=30)
        username_entry.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        username_entry.focus()
        
        ttk.Label(main_frame, text="Password:").grid(row=2, column=0, sticky=tk.W, pady=5)
        password_var = tk.StringVar()
        show_password_var = tk.BooleanVar(value=False)
        password_entry = ttk.Entry(main_frame, textvariable=password_var, show="●", width=30)
        password_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5)
        
        def toggle_password():
            if show_password_var.get():
                password_entry.config(show="●")
                show_password_var.set(False)
            else:
                password_entry.config(show="")
                show_password_var.set(True)
        
        eye_button = ttk.Button(main_frame, text="👁", width=3, command=toggle_password)
        eye_button.grid(row=2, column=2, sticky=tk.W, padx=(5, 0), pady=5)
        
        status_label = ttk.Label(main_frame, text="", font=('Arial', 9), foreground='red')
        status_label.grid(row=3, column=0, columnspan=3, pady=(5, 10))
        
        info_frame = ttk.Frame(main_frame)
        info_frame.grid(row=4, column=0, columnspan=3, pady=(0, 20))
        
        ttk.Label(info_frame, text="Enter your Wikimedia credentials", 
                 font=('Arial', 8), foreground='gray').pack()
        
        creds_frame = ttk.Frame(info_frame)
        creds_frame.pack()
        ttk.Label(creds_frame, text="Use ", font=('Arial', 8), foreground='gray').pack(side=tk.LEFT)
        
        def open_bot_password():
            import webbrowser
            webbrowser.open("https://commons.wikimedia.org/wiki/Special:BotPasswords")
        
        bot_link = ttk.Button(creds_frame, text="bot password", command=open_bot_password, 
                             style='Link.TButton', cursor='hand2')
        bot_link.pack(side=tk.LEFT)
        ttk.Label(creds_frame, text=" if you have 2FA enabled", 
                 font=('Arial', 8), foreground='gray').pack(side=tk.LEFT)
        
        def do_login():
            username = username_var.get().strip()
            password = password_var.get().strip()
            
            if not username or not password:
                status_label.config(text="Please enter both username and password", foreground='red')
                return
            
            status_label.config(text="Logging in...", foreground='blue')
            login_window.update()
            
            # Create config files
            if self.create_config_files(username, password):
                # Try to initialize pywikibot
                if self.test_login():
                    self.username = username
                    self.password = password
                    self.is_logged_in = True
                    family = self.family_var.get()
                    mylang = self.mylang_var.get()
                    self.login_status_var.set(f"Logged in as {username}")
                    self.login_status_label.config(foreground='green')
                    self.login_btn.config(text="Logout", command=self.do_logout)
                    self.log_message(f"Successfully logged in as {username} to {mylang}.{family}")
                    login_window.destroy()
                else:
                    status_label.config(text="Login failed! Please check credentials and try again.", foreground='red')
                    self.log_message("Login failed - invalid credentials or connection issue", "ERROR")
                    self.cleanup_config_files()
            else:
                status_label.config(text="Failed to create config files. Try again.", foreground='red')
        
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=5, column=0, columnspan=3)
        
        ttk.Button(button_frame, text="Login", command=do_login, width=15).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=login_window.destroy, width=15).pack(side=tk.LEFT, padx=5)
        
        login_window.bind('<Return>', lambda e: do_login())
        
        main_frame.columnconfigure(1, weight=1)
        login_window.columnconfigure(0, weight=1)
        login_window.rowconfigure(0, weight=1)
    
    def do_logout(self):
        """Logout and cleanup"""
        self.cleanup_config_files()
        self.username = None
        self.password = None
        self.is_logged_in = False
        self.login_status_var.set("Not logged in")
        self.login_status_label.config(foreground='red')
        self.login_btn.config(text="Login", command=self.show_login_window)
        self.log_message("Logged out successfully")
    
    def test_login(self):
        """Test login credentials"""
        try:
            self.log_message("Testing login credentials...")            
            # Set environment variables
            os.environ['PYWIKIBOT_DIR'] = CONFIG_DIR
            if CONFIG_DIR not in sys.path:
                sys.path.insert(0, CONFIG_DIR)
            # Force reload
            modules_to_remove = [key for key in sys.modules.keys() if key.startswith('pywikibot')]
            for module in modules_to_remove:
                del sys.modules[module]
            
            # Wait for files
            for _ in range(10):
                if os.path.exists(USER_CONFIG_PATH) and os.path.exists(PASSWORD_FILE_PATH):
                    try:
                        with open(PASSWORD_FILE_PATH, 'r', encoding='utf-8') as tf:
                            if tf.read(1) is not None:
                                break
                    except Exception:
                        pass
                time.sleep(0.1)
            
            import importlib
            pywikibot = importlib.import_module('pywikibot')
            
            family = self.family_var.get()
            mylang = self.mylang_var.get()
            
            test_site = pywikibot.Site(mylang, family)
            test_site.login()
            
            self.log_message("Login test successful")
            return True
            
        except Exception as e:
            self.log_message(f"Login test failed: {str(e)}", "ERROR")
            return False
    
    def create_config_files(self, username, password):
        """Create Pywikibot configuration files (use family and mylang from UI)"""
        try:
            self.log_message("Creating configuration files...")
            
            family = self.family_var.get()
            mylang = self.mylang_var.get()
            
            # Creating user-config.py
            user_config_content = f"""# -*- coding: utf-8 -*-
import sys
import os

family = '{family}'
mylang = '{mylang}'
usernames['{family}']['{mylang}'] = '{username}'
password_file = r"{PASSWORD_FILE_PATH}"
maxlag = 60
put_throttle = 1
console_encoding = 'utf-8'
max_retries = 10
simulate = False
textfile_encoding = 'utf-8'

# Suppress interactive password prompts
os.environ['PYWIKIBOT_NO_USER_CONFIG'] = '2'
"""
            with open(USER_CONFIG_PATH, 'w', encoding='utf-8') as f:
                f.write(user_config_content)
                f.flush()
                os.fsync(f.fileno())
            safe_chmod(USER_CONFIG_PATH, 0o600)

            # Creating user-password.py
            password_content = f"""# -*- coding: utf-8 -*-
# Password file for pywikibot
('{mylang}', '{family}', '{username}', '{password}')
"""
            with open(PASSWORD_FILE_PATH, 'w', encoding='utf-8') as f:
                f.write(password_content)
                f.flush()
                os.fsync(f.fileno())
            safe_chmod(PASSWORD_FILE_PATH, 0o600)
            
            time.sleep(0.5)
            
            self.log_message("Configuration files created successfully")
            return True

        except Exception as e:
            self.log_message(f"Failed to create config files: {str(e)}", "ERROR")
            return False
    
    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="Select input file",
            filetypes=[
                ("Supported files", "*.xlsx *.xls *.csv *.json"),
                ("Excel files", "*.xlsx *.xls"),
                ("CSV files", "*.csv"),
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        if filename:
            self.input_file.set(filename)
            # Setting output file to same directory of input
            input_dir = os.path.dirname(filename)
            input_basename = os.path.basename(filename)
            input_name, input_ext = os.path.splitext(input_basename)
            
            # Determine output extension based on input
            if input_ext.lower() in ['.csv']:
                output_ext = '.csv'
            elif input_ext.lower() in ['.json']:
                output_ext = '.json'
            else:
                output_ext = '.xlsx'
            
            output_filename = os.path.join(input_dir, f"{input_name}_results{output_ext}")
            self.output_file.set(output_filename)
            
    def read_input_file(self, filepath):
        """Read input file (Excel, CSV, or JSON) and return dataframe"""
        try:
            _, ext = os.path.splitext(filepath)
            ext = ext.lower()
            
            if ext in ['.xlsx', '.xls']:
                self.log_message(f"Reading Excel file: {filepath}")
                # Read Excel without evaluating formulas - treat everything as strings
                import openpyxl
                from openpyxl.utils.exceptions import InvalidFileException
                
                try:
                    # Use openpyxl to read raw cell values without formula evaluation
                    wb = openpyxl.load_workbook(filepath, data_only=False)
                    ws = wb.active
                    
                    rows = []
                    for row in ws.iter_rows(values_only=False):
                        row_data = []
                        for cell in row:
                            # Get the actual cell value, not the formula result
                            if cell.value is not None:
                                # If it's a formula, get the formula string without the =
                                if hasattr(cell, 'value') and isinstance(cell.value, str):
                                    row_data.append(cell.value)
                                else:
                                    row_data.append(str(cell.value) if cell.value is not None else '')
                            else:
                                row_data.append('')
                        rows.append(row_data)
                    
                    df = pd.DataFrame(rows)
                    self.log_message(f"Successfully read Excel with {len(df)} rows (raw values)")
                    
                except Exception as openpyxl_error:
                    self.log_message(f"openpyxl failed, trying pandas: {openpyxl_error}", "WARNING")
                    # Fallback to pandas
                    df = pd.read_excel(filepath, header=None, dtype=str)
            elif ext == '.csv':
                self.log_message(f"Reading CSV file: {filepath}")
                df = pd.read_csv(filepath, header=None)
            elif ext == '.json':
                self.log_message(f"Reading JSON file: {filepath}")
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                # Convert JSON to dataframe
                if isinstance(data, list):
                    # Assume list of objects with keys: file_path, target_filename, description
                    rows = []
                    for item in data:
                        if isinstance(item, dict):
                            rows.append([
                                item.get('file_path', ''),
                                item.get('target_filename', ''),
                                item.get('description', '')
                            ])
                        elif isinstance(item, list):
                            rows.append(item)
                    df = pd.DataFrame(rows)
                else:
                    self.log_message("Invalid JSON format - expected array", "ERROR")
                    return None
            else:
                self.log_message(f"Unsupported file type: {ext}", "ERROR")
                return None
            
            self.log_message(f"Successfully read {len(df)} rows from file")
            return df
            
        except Exception as e:
            self.log_message(f"Error reading file: {str(e)}", "ERROR")
            return None
    
    def log_message(self, message, level="INFO"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {level}: {message}\n"
        
        self.log_text.insert(tk.END, formatted_message)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
        if level == "ERROR":
            self.logger.error(message)
        elif level == "WARNING":
            self.logger.warning(message)
        else:
            self.logger.info(message)
            
    def update_internet_status(self, status):
        """Update internet status in UI"""
        self.internet_status.set(status)
        if status == "Active":
            self.internet_status_label.config(foreground="green")
        elif status == "Inactive":
            self.internet_status_label.config(foreground="red")
        else:
            self.internet_status_label.config(foreground="gray")
            
    def test_internet_connection(self, show_success=False):
        """Test internet connectivity using multiple reliable endpoints"""
        test_urls = [
            "https://www.google.com",
            "https://commons.wikimedia.org",
            "https://www.cloudflare.com",
            "https://8.8.8.8" 
        ]
        
        for url in test_urls:
            try:
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    self.update_internet_status("Active")
                    if show_success:
                        self.log_message("Internet connection is working")
                        messagebox.showinfo("Connection Test", "Internet connection is working!")
                    return True
            except:
                continue
                
        try:
            import socket
            socket.gethostbyname("google.com")
            self.update_internet_status("Active")
            return True
        except:
            pass
            
        self.update_internet_status("Inactive")
        if show_success:
            self.log_message("Internet connection failed", "ERROR")
        return False
            
    def clear_reset(self):
        """Clear all fields and reset the application state"""
        if self.is_running:
            response = messagebox.askyesno(
                "Upload in Progress",
                "An upload is currently in progress. Are you sure you want to reset?"
            )
            if not response:
                return
            # Stop the upload
            self.is_running = False
            self.is_paused = False
            self.stop_event.set()
            if self.executor:
                self.executor.shutdown(wait=False)
        
        # Logout if logged in
        if self.is_logged_in:
            self.do_logout()
        
        # Clear input fields
        self.input_file.set("")
        self.output_file.set("upload_results.xlsx")
        self.num_workers_var.set(1)
        self.pause_seconds_var.set(10)
        self.max_attempts_var.set(10)
        self.pause_after_upload_var.set(0.2)
        self.ignore_warnings_var.set("True")
        
        # Reset family and lang to defaults
        self.family_var.set("commons")
        self.mylang_var.set("commons")
        
        # Reset stop event for next run
        self.stop_event.clear()
        
        # Reset counters
        self.processed_files = 0
        self.successful_uploads = 0
        self.failed_uploads = 0
        self.total_files = 0
        self.results = []
        self.start_time = None
        
        # Reset UI elements
        self.progress_var.set(0)
        self.status_label.config(text="Ready")
        self.stats_label.config(text="Files: 0/0 | Success: 0 | Failed: 0")
        self.time_label.config(text="Time: 00:00:00 | ETA: --:--:--")
        self.update_internet_status("Unknown")
        
        # Clear log
        self.log_text.delete(1.0, tk.END)
        
        # Reset buttons
        self.start_btn.config(state=tk.NORMAL)
        self.pause_btn.config(state=tk.DISABLED, text="Pause")
        self.stop_btn.config(state=tk.DISABLED)
        
        self.log_message("Application reset successfully")
        messagebox.showinfo("Reset Complete", "All fields and settings have been reset")
            
    def download_file_from_url(self, url, max_retries):
        """Download file from URL with retries"""
        temp_file = None
        try:
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='_download')
            temp_path = temp_file.name
            temp_file.close()
            
            for attempt in range(max_retries):
                try:
                    self.log_message(f"Downloading from URL (attempt {attempt + 1}/{max_retries}): {url}")
                    response = requests.get(url, timeout=30, stream=True)
                    response.raise_for_status()
                    
                    with open(temp_path, 'wb') as f:
                        for chunk in response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    
                    # Check if file is not empty
                    if os.path.getsize(temp_path) > 0:
                        self.log_message(f"Successfully downloaded from URL: {url}")
                        return temp_path
                    else:
                        self.log_message(f"Downloaded file is empty (attempt {attempt + 1}/{max_retries})", "WARNING")
                        
                except Exception as e:
                    self.log_message(f"Download failed (attempt {attempt + 1}/{max_retries}): {str(e)}", "WARNING")
                    
                if attempt < max_retries - 1:
                    time.sleep(self.pause_seconds_var.get())
            
            # If all retries failed, try Wayback Machine
            self.log_message(f"All download attempts failed, trying Wayback Machine for: {url}")
            return self.download_from_wayback(url, max_retries)
            
        except Exception as e:
            self.log_message(f"Error in download_file_from_url: {str(e)}", "ERROR")
            if temp_file and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except:
                    pass
            return None
    
    def download_from_wayback(self, url, max_retries):
        """Download file from Wayback Machine (oldest snapshot)"""
        temp_file = None
        try:
            # Get oldest snapshot from Wayback Machine CDX API
            cdx_url = f"http://web.archive.org/cdx/search/cdx?url={url}&limit=1&sort=timestamp"
            
            for attempt in range(max_retries):
                try:
                    self.log_message(f"Querying Wayback Machine (attempt {attempt + 1}/{max_retries})")
                    cdx_response = requests.get(cdx_url, timeout=30)
                    cdx_response.raise_for_status()
                    
                    if not cdx_response.text.strip():
                        self.log_message("No snapshots found in Wayback Machine", "WARNING")
                        return None
                    
                    # Parse CDX response (space-separated)
                    cdx_line = cdx_response.text.strip().split('\n')[0]
                    parts = cdx_line.split()
                    
                    if len(parts) < 2:
                        self.log_message("Invalid CDX response from Wayback Machine", "WARNING")
                        return None
                    
                    timestamp = parts[1]
                    wayback_url = f"http://web.archive.org/web/{timestamp}id_/{url}"
                    
                    self.log_message(f"Found Wayback snapshot from {timestamp[:8]}, downloading...")
                    
                    # Download from Wayback Machine
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='_wayback')
                    temp_path = temp_file.name
                    temp_file.close()
                    
                    wb_response = requests.get(wayback_url, timeout=60, stream=True)
                    wb_response.raise_for_status()
                    
                    with open(temp_path, 'wb') as f:
                        for chunk in wb_response.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    
                    # Check if file is not empty
                    if os.path.getsize(temp_path) > 0:
                        self.log_message(f"Successfully downloaded from Wayback Machine")
                        return temp_path
                    else:
                        self.log_message(f"Wayback file is empty (attempt {attempt + 1}/{max_retries})", "WARNING")
                        os.remove(temp_path)
                        
                except Exception as e:
                    self.log_message(f"Wayback download failed (attempt {attempt + 1}/{max_retries}): {str(e)}", "WARNING")
                    if temp_file and os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                        except:
                            pass
                
                if attempt < max_retries - 1:
                    time.sleep(self.pause_seconds_var.get())
            
            self.log_message("All Wayback Machine attempts failed", "ERROR")
            return None
            
        except Exception as e:
            self.log_message(f"Error in download_from_wayback: {str(e)}", "ERROR")
            return None
    
    def is_youtube_url(self, url):
        """Check if URL is a YouTube video"""
        youtube_domains = ['youtube.com', 'youtu.be', 'm.youtube.com', 'www.youtube.com']
        try:
            parsed = urlparse(url)
            return any(domain in parsed.netloc for domain in youtube_domains)
        except:
            return False
    
    def download_youtube_video(self, url, max_retries):
        """Download YouTube video using yt-dlp with browser cookie fallback"""
        try:
            if not YT_DLP_AVAILABLE:
                self.log_message("yt-dlp not installed. Install with: pip install yt-dlp", "ERROR")
                return None

            # Build list of strategies to try in order:
            # 1. No cookies (anonymous, works for non-restricted videos)
            # 2. Chrome cookies
            # 3. Firefox cookies
            # 4. Edge cookies
            strategies = [
                {'name': 'anonymous',       'extra': {}},
                {'name': 'Chrome cookies',  'extra': {'cookiesfrombrowser': ('chrome',)}},
                {'name': 'Firefox cookies', 'extra': {'cookiesfrombrowser': ('firefox',)}},
                {'name': 'Edge cookies',    'extra': {'cookiesfrombrowser': ('edge',)}},
            ]

            self.log_message(f"Downloading YouTube video: {url}")

            for strategy in strategies:
                if self.stop_event.is_set():
                    self.log_message("YouTube download cancelled by user")
                    return None

                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='_youtube.mp4')
                temp_path = temp_file.name
                temp_file.close()
                os.remove(temp_path)

                for attempt in range(max_retries):
                    if self.stop_event.is_set():
                        self.log_message("YouTube download cancelled by user")
                        if os.path.exists(temp_path):
                            try:
                                os.remove(temp_path)
                            except:
                                pass
                        return None

                    self.log_message(f"YouTube download attempt {attempt + 1}/{max_retries} [{strategy['name']}]...")

                    # Try formats in order: direct mp4 > any non-HLS > any best
                    # proto filter avoids HLS/DASH streams that produce empty files without ffmpeg
                    format_options = [
                        'best[ext=mp4][protocol=https]',
                        'best[ext=mp4][protocol=http]',
                        'best[protocol=https]',
                        'best[protocol=http]',
                        'best',
                    ]
                    selected_format = format_options[min(attempt, len(format_options) - 1)]
                    ydl_opts = {
                        'format': selected_format,
                        'outtmpl': temp_path,
                        'quiet': True,
                        'no_warnings': True,
                    }
                    ydl_opts.update(strategy['extra'])

                    try:
                        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                            ydl.download([url])

                        if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
                            self.log_message(f"Successfully downloaded YouTube video [{strategy['name']}]: {temp_path}")
                            return temp_path
                        else:
                            self.log_message(f"Download produced empty file [{strategy['name']}] (attempt {attempt + 1}/{max_retries})", "WARNING")

                    except Exception as e:
                        err = str(e)
                        self.log_message(f"Download error [{strategy['name']}] attempt {attempt + 1}/{max_retries}: {err}", "WARNING")
                        # If bot-detection error, no point retrying same strategy
                        if 'Sign in to confirm' in err or 'not a bot' in err:
                            self.log_message(f"Bot detection hit with [{strategy['name']}], trying next strategy...")
                            break
                    # If empty file, move to next format selector on next attempt (already handled above)
                    # but if we've exhausted all format options, break to next strategy
                    if os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                        except:
                            pass
                    if attempt >= len(format_options) - 1 and not (os.path.exists(temp_path) and os.path.getsize(temp_path) > 0):
                        self.log_message(f"All format selectors exhausted for [{strategy['name']}], trying next strategy...")
                        break

                    # Cleanup failed attempt file
                    if os.path.exists(temp_path):
                        try:
                            os.remove(temp_path)
                        except:
                            pass

                    if attempt < max_retries - 1:
                        self.log_message(f"Waiting {self.pause_seconds_var.get()} seconds before retry...")
                        time.sleep(self.pause_seconds_var.get())

                # Cleanup between strategies
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass

            self.log_message("All YouTube download strategies failed. Open a browser, log into YouTube, then retry.", "ERROR")
            return None

        except Exception as e:
            self.log_message(f"Error in download_youtube_video: {str(e)}", "ERROR")
            return None
    
    def wait_for_internet(self):
        """Wait for internet connection to be restored"""
        retry_count = 0
        while self.is_running and not self.test_internet_connection():
            retry_count += 1
            if retry_count == 1:  # Only log on first attempt
                self.log_message("Waiting for internet connection...")
            time.sleep(10)
        
        if self.is_running and retry_count > 0:
            self.log_message("Internet connection restored")
        return self.is_running
        
    def initialize_pywikibot(self):
        """Initialize Pywikibot with proper config paths"""
        try:
            # Set environment variables for pywikibot
            os.environ['PYWIKIBOT_DIR'] = CONFIG_DIR
            # Add CONFIG_DIR to sys.path so pywikibot can find user-config.py
            if CONFIG_DIR not in sys.path:
                sys.path.insert(0, CONFIG_DIR)
            # Force reload of pywikibot to use new config
            modules_to_remove = [key for key in sys.modules.keys() if key.startswith('pywikibot')]
            for module in modules_to_remove:
                del sys.modules[module]

            # Wait for config files to be readable
            for _ in range(10):  # up to ~1 second total (10 * 0.1s)
                if os.path.exists(USER_CONFIG_PATH) and os.path.exists(PASSWORD_FILE_PATH):
                    try:
                        # quick open test to ensure filesystem returns readable content
                        with open(PASSWORD_FILE_PATH, 'r', encoding='utf-8') as tf:
                            if tf.read(1) is not None:
                                break
                    except Exception:
                        pass
                time.sleep(0.1)
            else:
                self.log_message("Config files not ready for pywikibot.", "ERROR")
                return False
            
            import importlib
            pywikibot = importlib.import_module('pywikibot')
            from pywikibot import FilePage as _FilePage

            self.pywikibot = pywikibot
            self.FilePage = _FilePage

            # Check if config files exist
            if not os.path.exists(USER_CONFIG_PATH) or not os.path.exists(PASSWORD_FILE_PATH):
                self.log_message(f"Config files missing: {USER_CONFIG_PATH}, {PASSWORD_FILE_PATH}", "ERROR")
                return False
                
            # Verify password file is readable
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    with open(PASSWORD_FILE_PATH, 'r', encoding='utf-8') as f:
                        content = f.read()
                        if content.strip():  # Ensure file has content
                            break
                except Exception as e:
                    if attempt < max_retries - 1:
                        time.sleep(0.5)
                    else:
                        self.log_message(f"Cannot read password file after {max_retries} attempts: {e}", "ERROR")
                        return False
            
            family = self.family_var.get()
            mylang = self.mylang_var.get()
            
            self.site = self.pywikibot.Site(mylang, family)
            self.site.login()

            # Debug info
            self.log_message(f"Config directory: {CONFIG_DIR}")
            if getattr(sys.modules[__name__], '__compiled__', False):
                self.log_message(f"Pywikibot data directory: {PYWIKIBOT_DATA_DIR}")
                self.log_message(f"Running as compiled executable")
            self.log_message(f"Successfully logged in as {self.username} in {mylang} {family}")
            return True

        except Exception as e:
            self.log_message(f"Failed to initialize Pywikibot: {str(e)}", "ERROR")
            import traceback
            self.log_message(f"Traceback: {traceback.format_exc()}", "ERROR")
            return False

    def convert_video_to_webm(self, input_path, max_retries=3):
        """Convert video file to WebM format using moviepy"""
        try:
            if not MOVIEPY_AVAILABLE:
                self.log_message("moviepy not installed. Install with: pip install moviepy", "ERROR")
                return None
            
            # Create temporary output file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.webm')
            output_path = temp_file.name
            temp_file.close()
            
            self.log_message(f"Converting video to WebM: {input_path}")
            
            for attempt in range(max_retries):
                if self.stop_event.is_set():
                    self.log_message("Video conversion cancelled by user")
                    return None
                try:
                    # Load video
                    self.log_message(f"Loading video file (attempt {attempt + 1}/{max_retries})...")
                    video = VideoFileClip(input_path)
                    
                    # Convert to WebM with good quality settings
                    # Using libvpx-vp9 codec (VP9) which is preferred for Wikimedia Commons
                    self.log_message(f"Converting to WebM format...")
                    # Handle videos without fps info
                    fps_value = video.fps if video.fps and video.fps > 0 else 30
                    # Create a temp audio file with .opus extension so moviepy
                    # can resolve the libopus codec correctly
                    temp_audio = tempfile.NamedTemporaryFile(delete=False, suffix='.opus')
                    temp_audio_path = temp_audio.name
                    temp_audio.close()
                    try:
                        video.write_videofile(
                            output_path,
                            codec='libvpx-vp9',
                            audio_codec='libopus',
                            temp_audiofile=temp_audio_path,
                            bitrate='2000k',
                            audio_bitrate='128k',
                            audio_fps=48000,
                            fps=fps_value,
                            threads=4,
                            logger=None
                        )
                    except Exception as opus_err:
                        if 'unknown' in str(opus_err).lower() or 'audio_codec' in str(opus_err).lower():
                            # libopus not available in this ffmpeg build, fall back to libvorbis
                            self.log_message("libopus unavailable, retrying with libvorbis...", "WARNING")
                            temp_audio_ogg = tempfile.NamedTemporaryFile(delete=False, suffix='.ogg')
                            temp_audio_ogg_path = temp_audio_ogg.name
                            temp_audio_ogg.close()
                            try:
                                video.write_videofile(
                                    output_path,
                                    codec='libvpx-vp9',
                                    audio_codec='libvorbis',
                                    temp_audiofile=temp_audio_ogg_path,
                                    bitrate='2000k',
                                    audio_bitrate='128k',
                                    audio_fps=48000,
                                    fps=fps_value,
                                    threads=4,
                                    logger=None
                                )
                            finally:
                                try:
                                    os.remove(temp_audio_ogg_path)
                                except:
                                    pass
                        else:
                            raise
                    finally:
                        try:
                            os.remove(temp_audio_path)
                        except:
                            pass
                    
                    # Close the video to free resources
                    video.close()
                    
                    # Verify output file exists and is not empty
                    if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                        self.log_message(f"Successfully converted video to WebM: {output_path}")
                        return output_path
                    else:
                        self.log_message(f"Conversion produced empty file (attempt {attempt + 1}/{max_retries})", "WARNING")
                        
                except Exception as e:
                    self.log_message(f"Conversion error (attempt {attempt + 1}/{max_retries}): {str(e)}", "WARNING")
                    try:
                        video.close()
                    except:
                        pass
                
                # Cleanup failed attempt
                if os.path.exists(output_path):
                    try:
                        os.remove(output_path)
                    except:
                        pass
                
                if attempt < max_retries - 1:
                    self.log_message(f"Waiting 2 seconds before retry...")
                    time.sleep(2)
            
            self.log_message("All video conversion attempts failed", "ERROR")
            return None
            
        except Exception as e:
            self.log_message(f"Error in convert_video_to_webm: {str(e)}", "ERROR")
            try:
                if 'video' in locals():
                    video.close()
            except:
                pass
            return None
    
    def get_extension_from_file(self, file_path):
        """Get extension from actual file content using file signatures and PIL"""
        try:
            # Read file signatures (magic numbers) - need more bytes for some formats
            with open(file_path, 'rb') as f:
                header = f.read(512)  # Read more bytes for better detection            
            # Check file signatures (magic numbers)
            if header[:4] == b'\x89PNG':
                return '.png'
            elif header[:3] == b'\xff\xd8\xff':
                return '.jpg'
            elif header[:6] in (b'GIF87a', b'GIF89a'):
                return '.gif'
            elif header[:4] == b'\x89PNG':
                with open(file_path, 'rb') as f:
                    content = f.read(4096)
                if b'acTL' in content:
                    return '.apng'
                return '.png'
            elif header[:4] == b'RIFF' and header[8:12] == b'WEBP':
                return '.webp'
            elif header[:4] == b'%PDF':
                return '.pdf'
            elif header[:4] == b'\x1a\x45\xdf\xa3':
                if b'webm' in header[:100].lower() or b'matroska' in header[:100].lower():
                    _, ext = os.path.splitext(file_path)
                    if ext.lower() == '.webm':
                        return '.webm'
                return '.webm'
            
            elif header[:4] == b'OggS':
                _, ext = os.path.splitext(file_path)
                if ext.lower() in ['.ogg', '.ogv', '.oga', '.ogx']:
                    return ext.lower()
                return '.ogg'
            
            elif b'<svg' in header[:512].lower():
                return '.svg'
            elif header[:5] == b'<?xml':
                with open(file_path, 'rb') as f:
                    content = f.read(2048)
                if b'<svg' in content.lower():
                    return '.svg'
            
            elif header[:4] == b'MThd':
                return '.mid'
            
            elif header[:8] == b'AT&TFORM' or header[4:8] == b'DJVU' or header[4:8] == b'DJVM':
                return '.djvu'
            
            elif header[:9] == b'gimp xcf ':
                return '.xcf'
            
            elif header[:4] in (b'II*\x00', b'MM\x00*'):
                return '.tif'
            
            elif header[:4] == b'RIFF' and header[8:12] == b'WAVE':
                return '.wav'
            
            elif header[:4] == b'fLaC':
                return '.flac'
            
            elif header[:3] == b'ID3' or (header[0] == 0xFF and (header[1] & 0xE0) == 0xE0):
                return '.mp3'
            
            elif header[:4] == b'OggS':
                with open(file_path, 'rb') as f:
                    content = f.read(4096)
                if b'OpusHead' in content:
                    return '.opus'
                _, ext = os.path.splitext(file_path)
                if ext.lower() in ['.ogg', '.ogv', '.oga', '.ogx', '.opus']:
                    return ext.lower()
                return '.ogg'
            
            elif header[:5] == b'solid' or header[:80].startswith(b'solid'):
                return '.stl'
            elif len(header) >= 84:
                try:
                    import struct
                    triangle_count = struct.unpack('<I', header[80:84])[0]
                    if 0 < triangle_count < 100000000:
                        return '.stl'
                except:
                    pass
            elif header[:2] == b'BM':
                return '.bmp'
            
            # For remaining image files, use PIL as fallback
            try:
                with Image.open(file_path) as img:
                    format_ext = img.format.lower()
                    if format_ext == 'jpeg':
                        return '.jpg'
                    elif format_ext == 'tiff':
                        return '.tif'
                    return f'.{format_ext}'
            except:
                pass
            
            # Last resort: use file extension from the actual file path (not target name)
            # This helps with formats we can't detect from magic numbers
            # Use rsplit to get extension after the LAST dot only
            if '.' in file_path:
                ext = '.' + file_path.rsplit('.', 1)[-1].lower()
                if ext in ALLOWED_EXTENSIONS:
                    self.log_message(f"Using file extension from path for {file_path}: {ext}", "INFO")
                    return ext
                
            return None
                
        except Exception as e:
            self.log_message(f"Could not determine file type for {file_path}: {e}", "WARNING")
            return None

    def verify_upload(self, original_file_path, target_filename, expected_wikitext, file_page):
        """Verify uploaded file matches original"""
        try:
            # Handle auto-incremented filenames
            actual_filename = file_page.title(with_ns=False)
            self.log_message(f"Verifying upload: {actual_filename}")
            
            # Get file info from Commons
            try:
                # Reload page to get latest info
                file_page = self.FilePage(self.site, f'File:{actual_filename}')
                file_info = file_page.latest_file_info
                uploaded_size = file_info.size
            except Exception as e:
                return f"Not OK: Could not get file info - {str(e)}"
            
            # Get original file size
            try:
                original_size = os.path.getsize(original_file_path)
            except Exception as e:
                return f"Not OK: Could not get original file size - {str(e)}"
            
            # Compare file sizes (allow 2 byte difference in either direction)
            size_diff = abs(uploaded_size - original_size)
            if size_diff > 2:
                return f"Not OK: Size mismatch (original: {original_size}, uploaded: {uploaded_size}, diff: {size_diff})"
            
            # Verify wikitext content
            try:
                uploaded_text = file_page.text
                # Remove category added by pypan for comparison
                expected_clean = expected_wikitext.replace("\n[[Category: Uploaded with pypan]]", "")
                uploaded_clean = uploaded_text.replace("[[Category: Uploaded with pypan]]", "")
                
                if expected_clean.strip() == uploaded_clean.strip():
                    return "Verified"
                else:
                    return "Not OK: Wikitext mismatch"
            except Exception as e:
                return f"Not OK: Could not verify wikitext - {str(e)}"
                
        except Exception as e:
            return f"Not OK: Verification error - {str(e)}"
        
    def upload_single_file(self, row_data, row_index):
        """Upload a single file with retry logic"""
        from pywikibot.exceptions import UploadError
        file_path, target_filename, description = row_data
        
        # Check if file_path is a URL
        is_url = False
        downloaded_file = None
        converted_file = None
        try:
            parsed = urlparse(file_path)
            is_url = parsed.scheme in ('http', 'https')
        except:
            is_url = False
        
        # Download file if it's a URL
        if is_url:
            self.log_message(f"Detected URL: {file_path}")
            
            # Check if it's a YouTube URL
            if self.is_youtube_url(file_path):
                self.log_message("YouTube URL detected, using yt-dlp...")
                downloaded_file = self.download_youtube_video(file_path, self.max_attempts_var.get())
            else:
                downloaded_file = self.download_file_from_url(file_path, self.max_attempts_var.get())
            
            if not downloaded_file:
                error_msg = 'Could not download YouTube video' if self.is_youtube_url(file_path) else 'Could not download file from URL or Wayback Machine'
                result = {
                    'row': row_index + 1,
                    'file_path': file_path,
                    'target_filename': target_filename,
                    'status': 'Failed',
                    'error': error_msg,
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                return result
            
            file_path = downloaded_file
        
        # Check if file exists first
        if not os.path.exists(file_path):
            result = {
                'row': row_index + 1,
                'file_path': file_path,
                'target_filename': target_filename,
                'status': 'Skipped',
                'error': 'File not found',
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            return result
        
        # Check if file needs video conversion
        _, original_ext = os.path.splitext(file_path)
        if original_ext.lower() in VIDEO_FORMATS_TO_CONVERT:
            self.log_message(f"Detected video format {original_ext}, converting to WebM...")
            converted_file = self.convert_video_to_webm(file_path)
            
            if not converted_file:
                result = {
                    'row': row_index + 1,
                    'file_path': file_path,
                    'target_filename': target_filename,
                    'status': 'Failed',
                    'error': f'Could not convert {original_ext} to WebM. Install moviepy: pip install moviepy',
                    'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                }
                # Cleanup downloaded file if exists
                if downloaded_file and os.path.exists(downloaded_file):
                    try:
                        os.remove(downloaded_file)
                        self.log_message(f"Cleaned up downloaded temp file: {downloaded_file}")
                    except Exception as e:
                        self.log_message(f"Could not remove downloaded file {downloaded_file}: {e}", "WARNING")
                return result
            
            file_path = converted_file
            # Update target filename to use .webm extension
            if '.' in target_filename:
                target_filename = target_filename.rsplit('.', 1)[0] + '.webm'
            else:
                target_filename = target_filename + '.webm'
            self.log_message(f"Video converted successfully, new filename: {target_filename}")
        
        actual_file_ext = self.get_extension_from_file(file_path)
        
        # Skip if no extension could be determined
        if not actual_file_ext:
            # Cleanup temp files
            if downloaded_file and os.path.exists(downloaded_file):
                try:
                    os.remove(downloaded_file)
                except:
                    pass
            if converted_file and os.path.exists(converted_file):
                try:
                    os.remove(converted_file)
                except:
                    pass
            result = {
                'row': row_index + 1,
                'file_path': file_path,
                'target_filename': target_filename,
                'status': 'Skipped',
                'error': 'Could not determine file extension',
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            self.log_message(f"Skipping {file_path}: Could not determine file extension", "WARNING")
            return result
        
        # Check if extension is allowed
        if actual_file_ext not in ALLOWED_EXTENSIONS:
            # Cleanup temp files
            if downloaded_file and os.path.exists(downloaded_file):
                try:
                    os.remove(downloaded_file)
                except:
                    pass
            if converted_file and os.path.exists(converted_file):
                try:
                    os.remove(converted_file)
                except:
                    pass
            result = {
                'row': row_index + 1,
                'file_path': file_path,
                'target_filename': target_filename,
                'status': 'Skipped',
                'error': f'File extension {actual_file_ext} not allowed',
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            self.log_message(f"Skipping {file_path}: Extension {actual_file_ext} not in allowed list", "WARNING")
            return result
        
        # Remove any existing extension from target filename and add the correct one
        target_filename_base = target_filename
        if '.' in target_filename:
            parts = target_filename.rsplit('.', 1)
            if len(parts) == 2 and f'.{parts[1].lower()}' in ALLOWED_EXTENSIONS:
                target_filename_base = parts[0]
        
        # Sanitize filename to remove illegal characters
        target_filename_base = sanitize_filename(target_filename_base)
        target_filename = target_filename_base + actual_file_ext
                
        result = {
            'row': row_index + 1,
            'file_path': file_path,
            'target_filename': target_filename,
            'status': 'Failed',
            'error': '',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        for attempt in range(self.max_attempts_var.get()):
            try:
                # Check if paused
                while self.is_paused and self.is_running:
                    time.sleep(0.5)
                    
                if not self.is_running:
                    result['error'] = 'Upload stopped by user'
                    # Cleanup temp files
                    if downloaded_file and os.path.exists(downloaded_file):
                        try:
                            os.remove(downloaded_file)
                        except:
                            pass
                    if converted_file and os.path.exists(converted_file):
                        try:
                            os.remove(converted_file)
                        except:
                            pass
                    return result
                
                # Check internet connection before each upload
                if not self.test_internet_connection():
                    self.log_message(f"No internet connection for {target_filename}, waiting...", "WARNING")
                    if not self.wait_for_internet():
                        result['error'] = 'No internet connection'
                        return result
                
                # Check if file exists
                if not os.path.exists(file_path):
                    result['error'] = f'File not found: {file_path}'
                    return result
                
                # Create FilePage with ignore_extension to prevent validation issues with dots in filename
                original_target_filename = target_filename
                counter = 0
                file_page = self.FilePage(self.site, f'File:{target_filename}', ignore_extension=True)
                
                # Check if file already exists and auto-increment
                while file_page.exists():
                    counter += 1
                    # Split filename and extension
                    name_parts = original_target_filename.rsplit('.', 1)
                    if len(name_parts) == 2:
                        target_filename = f"{name_parts[0]} ({counter}).{name_parts[1]}"
                    else:
                        target_filename = f"{original_target_filename} ({counter})"
                    file_page = self.FilePage(self.site, f'File:{target_filename}', ignore_extension=True)
                    self.log_message(f"File exists, trying: {target_filename}")
                
                if counter > 0:
                    self.log_message(f"Using filename: {target_filename} (original was taken)")
                
                # Upload file
                self.log_message(f"Uploading {target_filename} (attempt {attempt + 1})")
                
                success = file_page.upload(
                    source=file_path,
                    comment=f"Pypan 0.2.1a0",
                    text=description,
                    ignore_warnings=(self.ignore_warnings_var.get() == "True")
                )
                
                if success:
                    result['status'] = 'Success'
                    result['error'] = ''
                    self.log_message(f"Successfully uploaded {target_filename}")
                    
                    # Verify upload 
                    verification_result = self.verify_upload(file_path, target_filename, description, file_page)
                    result['verification'] = verification_result
                    self.log_message(f"Verification: {verification_result}")
                    
                    # Update results file incrementally
                    self.results.append(result)
                    self.save_results()
                    self.results.pop()  # Remove to avoid duplicate when returned
                    
                    # Wait after successful upload
                    time.sleep(self.pause_after_upload_var.get())
                    return result
                else:
                    result['error'] = 'Upload failed - server response'
                    
            except UploadError as e:
                result['error'] = f'Upload warning: {str(e)}'
                self.log_message(f"Upload warning for {target_filename}: {str(e)}", "WARNING")
                
            except Exception as e:
                result['error'] = f'Exception: {str(e)}'
                self.log_message(f"Error uploading {target_filename}: {str(e)}", "ERROR")
            
            # Wait before retry 
            if attempt < self.max_attempts_var.get() - 1:
                self.log_message(f"Waiting {self.pause_seconds_var.get()} seconds before retry (attempt {attempt + 1}/{self.max_attempts_var.get()})")
                time.sleep(self.pause_seconds_var.get())
        
        # Cleanup downloaded file if it was from URL
        if downloaded_file and os.path.exists(downloaded_file):
            try:
                os.remove(downloaded_file)
                self.log_message(f"Cleaned up temporary file: {downloaded_file}")
            except Exception as e:
                self.log_message(f"Could not remove temporary file {downloaded_file}: {e}", "WARNING")
        
        # Cleanup converted file if it was created
        if converted_file and os.path.exists(converted_file):
            try:
                os.remove(converted_file)
                self.log_message(f"Cleaned up converted file: {converted_file}")
            except Exception as e:
                self.log_message(f"Could not remove converted file {converted_file}: {e}", "WARNING")
            
        return result
        
    def update_progress(self):
        """Update progress indicators"""
        if self.total_files > 0:
            progress = (self.processed_files / self.total_files) * 100
            self.progress_var.set(progress)
            
            # Update stats
            self.stats_label.config(
                text=f"Files: {self.processed_files}/{self.total_files} | "
                     f"Success: {self.successful_uploads} | "
                     f"Failed: {self.failed_uploads}"
            )
            
            # Calculate time and ETA
            if self.start_time:
                elapsed = time.time() - self.start_time
                elapsed_str = time.strftime("%H:%M:%S", time.gmtime(elapsed))
                
                if self.processed_files > 0:
                    avg_time_per_file = elapsed / self.processed_files
                    remaining_files = self.total_files - self.processed_files
                    eta_seconds = avg_time_per_file * remaining_files
                    eta_str = time.strftime("%H:%M:%S", time.gmtime(eta_seconds))
                else:
                    eta_str = "--:--:--"
                    
                self.time_label.config(text=f"Time: {elapsed_str} | ETA: {eta_str}")
                
    def save_results(self):
        """Save results to file (Excel, CSV, or JSON) with status in last column"""
        try:
            output_path = self.output_file.get()
            _, ext = os.path.splitext(output_path)
            ext = ext.lower()
            
            # Check if output file already exists and create unique name
            if os.path.exists(output_path):
                base_name, extension = os.path.splitext(output_path)
                counter = 1
                while os.path.exists(f"{base_name}_{counter}{extension}"):
                    counter += 1
                output_path = f"{base_name}_{counter}{extension}"
                self.output_file.set(output_path)
                self.log_message(f"Output file exists, using: {output_path}")
            
            input_df = self.read_input_file(self.input_file.get())
            if input_df is None:
                self.log_message("Could not read input file for results", "ERROR")
                return
            
            input_df['Upload_Status'] = ''
            input_df['Verification'] = ''
            
            # Update status for each row based on results
            for result in self.results:
                row_idx = result['row'] - 1  # Convert to 0-based index
                if row_idx < len(input_df):
                    if result['status'] == 'Success':
                        input_df.loc[row_idx, 'Upload_Status'] = 'Success'
                        input_df.loc[row_idx, 'Verification'] = result.get('verification', '')
                    elif result['status'] == 'Skipped':
                        input_df.loc[row_idx, 'Upload_Status'] = f"Skipped: {result['error']}"
                    else:
                        input_df.loc[row_idx, 'Upload_Status'] = f"Failed: {result['error']}"
            
            # Save based on file type
            if ext in ['.xlsx', '.xls']:
                # Use openpyxl directly to preserve formulas
                import openpyxl
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                
                # Write data row by row, preserving formulas
                for r_idx, row in input_df.iterrows():
                    for c_idx, value in enumerate(row):
                        if pd.notna(value):
                            str_value = str(value)
                            # Check if this is intended as text that starts with = (wikitext)
                            # vs an actual Excel formula
                            # For the description column (index 2), always treat as text
                            if c_idx == 2 and str_value.startswith('='):
                                # Escape with single quote to prevent Excel from treating as formula
                                ws.cell(row=r_idx + 1, column=c_idx + 1).value = "'" + str_value
                            elif str_value.startswith('='):
                                # For other columns, write as formula
                                ws.cell(row=r_idx + 1, column=c_idx + 1).value = str_value
                            else:
                                ws.cell(row=r_idx + 1, column=c_idx + 1).value = str_value
                        else:
                            ws.cell(row=r_idx + 1, column=c_idx + 1).value = ''
                wb.save(output_path)
            elif ext == '.csv':
                input_df.to_csv(output_path, index=False, header=False)
            elif ext == '.json':
                # Convert to JSON format
                json_data = []
                for _, row in input_df.iterrows():
                    json_data.append({
                        'file_path': str(row[0]) if pd.notna(row[0]) else '',
                        'target_filename': str(row[1]) if pd.notna(row[1]) else '',
                        'description': str(row[2]) if pd.notna(row[2]) else '',
                        'upload_status': str(row['Upload_Status']) if pd.notna(row['Upload_Status']) else '',
                        'verification': str(row['Verification']) if pd.notna(row['Verification']) else ''
                    })
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(json_data, f, indent=2, ensure_ascii=False)
            
            self.log_message(f"Results saved to {output_path}")
        except Exception as e:
            self.log_message(f"Error saving results: {str(e)}", "ERROR")
            
    def upload_worker_thread(self):
        """Main upload worker thread"""
        try:
            df = self.read_input_file(self.input_file.get())
            if df is None:
                self.log_message("Could not read input file", "ERROR")
                return
            
            self.total_files = len(df)
            
            if self.total_files == 0:
                self.log_message("No files found in input file", "ERROR")
                return
                
            self.log_message(f"Found {self.total_files} files to upload")
            
            if not self.initialize_pywikibot():
                return
                
            # Process files with thread pool
            with ThreadPoolExecutor(max_workers=self.num_workers_var.get()) as executor:
                self.executor = executor
                
                # Submit all tasks
                future_to_row = {}
                for index, row in df.iterrows():
                    if not self.is_running:
                        break
                        
                    file_path = str(row[0]) if pd.notna(row[0]) else ""
                    target_filename = str(row[1]) if pd.notna(row[1]) else ""
                    # Handle Excel formulas - if description starts with =, Excel might treat it as formula
                    # We need to read it as raw string
                    if pd.notna(row[2]):
                        description = str(row[2])
                        # If Excel stripped the leading =, try to detect and restore it
                        if not description.startswith('=') and not description.startswith('{'):
                            # Check if it looks like it should have started with =
                            if description.startswith('={{') or description.startswith('{int:'):
                                description = '=' + description
                    else:
                        description = ""
                    description += "\n[[Category: Uploaded with pypan]]"
                    
                    if file_path and target_filename:
                        # Log the description being used
                        self.log_message(f"Row {index + 1}: Using description (first 100 chars): {description[:100]}")
                        future = executor.submit(
                            self.upload_single_file, 
                            (file_path, target_filename, description),
                            index
                        )
                        future_to_row[future] = index
                        
                # Process completed tasks
                for future in as_completed(future_to_row):
                    if not self.is_running:
                        break
                        
                    try:
                        result = future.result()
                        self.results.append(result)
                        
                        self.processed_files += 1
                        
                        if result['status'] == 'Success':
                            self.successful_uploads += 1
                        else:
                            self.failed_uploads += 1
                            
                        self.update_progress()
                        
                    except Exception as e:
                        self.log_message(f"Error processing result: {str(e)}", "ERROR")
                        self.failed_uploads += 1
                        
            self.save_results()
            
        except Exception as e:
            self.log_message(f"Upload thread error: {str(e)}", "ERROR")
        finally:
            self.upload_finished()
            
    def upload_finished(self):
        """Called when upload process is finished"""
        self.is_running = False
        self.is_paused = False
        
        self.start_btn.config(state=tk.NORMAL)
        self.pause_btn.config(state=tk.DISABLED, text="Pause")
        self.stop_btn.config(state=tk.DISABLED)
        
        self.status_label.config(text="Upload completed")
        
        self.log_message("Upload process completed")
        self.log_message(f"Total: {self.total_files}, Success: {self.successful_uploads}, Failed: {self.failed_uploads}")
        
        # completion dialog
        messagebox.showinfo(
            "Upload Complete",
            f"Batch upload completed!\n\n"
            f"Total files: {self.total_files}\n"
            f"Successful: {self.successful_uploads}\n"
            f"Failed: {self.failed_uploads}\n\n"
            f"Results saved to:\n{self.output_file.get()}"
        )
        
        # Cleanup config files
        self.cleanup_config_files()
        
    def start_upload(self):
        """Start the upload process"""
        # Check if logged in
        if not self.is_logged_in:
            response = messagebox.askyesno(
                "Login Required",
                "You must be logged in to upload files. Would you like to login now?"
            )
            if response:
                self.show_login_window()
            return
        
        if not self.input_file.get():
            messagebox.showerror("Error", "Please select an input file")
            return
            
        if not os.path.exists(self.input_file.get()):
            messagebox.showerror("Error", "Input file does not exist")
            return
            
        # Reset counters and stop flag
        self.stop_event.clear()
        self.processed_files = 0
        self.successful_uploads = 0
        self.failed_uploads = 0
        self.results = []
        self.start_time = time.time()
        
        self.is_running = True
        self.is_paused = False
        self.start_btn.config(state=tk.DISABLED)
        self.pause_btn.config(state=tk.NORMAL, text="Pause")
        self.stop_btn.config(state=tk.NORMAL)
        
        self.status_label.config(text="Starting upload...")
        self.log_message("Starting upload process")
        
        # Start worker thread
        threading.Thread(target=self.upload_worker_thread, daemon=True).start()
        
    def pause_upload(self):
        """Pause/resume the upload process"""
        if self.is_paused:
            self.is_paused = False
            self.pause_btn.config(text="Pause")
            self.status_label.config(text="Resuming upload...")
            self.log_message("Upload resumed")
        else:
            self.is_paused = True
            self.pause_btn.config(text="Resume")
            self.status_label.config(text="Upload paused")
            self.log_message("Upload paused")
            
    def stop_upload(self):
        """Stop the upload process"""
        self.is_running = False
        self.is_paused = False
        self.stop_event.set()
        
        if self.executor:
            self.executor.shutdown(wait=False)
            
        self.status_label.config(text="Stopping upload...")
        self.log_message("Upload stopped by user")

def main():
    # Create config directory if running as compiled executable
    if getattr(sys.modules[__name__], '__compiled__', False):
        os.makedirs(CONFIG_DIR, exist_ok=True)
    
    root = tk.Tk()
    app = PyPan(root)
    root.mainloop()


if __name__ == "__main__":
    main()
